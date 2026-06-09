"""Append parsed ODS rows to the Pending sheet of Order.xlsx."""

from __future__ import annotations

import datetime as dt
import shutil
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ods_parser import OdsRow

PENDING_SHEET = "Pending"
FIRST_DATA_ROW = 4  # rows 1-3 are header/title/sum
LAST_FORMAT_COL = 7  # copy formatting for columns A-G (G stays empty but needs borders)
DATE_FORMAT = "m/d/yyyy"


def _copy_row_format(ws, template_row: int, target_row: int) -> None:
    """Replicate the cell formatting of template_row onto target_row (cols A-F).

    Mirrors the owner's manual "format painter" step: take the last existing row's
    look (font, fill, border, alignment, number format) and apply it to new rows.
    """
    for col in range(1, LAST_FORMAT_COL + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format


@dataclass
class OrderBatch:
    order_number: str
    date: dt.date
    rows: list[OdsRow]


def _existing_order_numbers(ws) -> set[str]:
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, min_col=5, max_col=5, values_only=True):
        val = row[0]
        if val:
            seen.add(str(val).strip())
    return seen


def _next_empty_row(ws) -> int:
    n = FIRST_DATA_ROW
    while ws.cell(row=n, column=1).value not in (None, ""):
        n += 1
    return n


def write_orders(
    master_xlsx: Path,
    batches: list[OrderBatch],
    backups_dir: Path,
) -> tuple[Path | None, int, int]:
    """Append batches to the master workbook IN PLACE.

    Before overwriting, a pristine copy of the master is saved to backups_dir as a
    restore point. Returns (backup_path, rows_added, orders_added); backup_path is
    None when nothing was appended (master left untouched).

    Skips any batch whose order_number is already present in column E.
    """
    wb = load_workbook(master_xlsx)
    if PENDING_SHEET not in wb.sheetnames:
        raise ValueError(f"{master_xlsx} has no '{PENDING_SHEET}' sheet")
    ws = wb[PENDING_SHEET]

    seen = _existing_order_numbers(ws)
    next_row = _next_empty_row(ws)
    # Template for formatting: the last existing data row. None if the sheet is empty.
    template_row = next_row - 1 if next_row - 1 >= FIRST_DATA_ROW else None

    rows_added = 0
    orders_added = 0
    for batch in batches:
        if batch.order_number in seen:
            continue
        if not batch.rows:
            continue
        for r in batch.rows:
            ws.cell(row=next_row, column=1, value=r.item_no)
            ws.cell(row=next_row, column=2, value=r.name)
            ws.cell(
                row=next_row,
                column=3,
                value=f"=VLOOKUP(A{next_row},SizeData!A$1:B$3974,2,FALSE)",
            )
            ws.cell(row=next_row, column=4, value=r.reserved_qty)
            ws.cell(row=next_row, column=5, value=batch.order_number)
            ws.cell(row=next_row, column=6, value=batch.date)
            if template_row is not None:
                _copy_row_format(ws, template_row, next_row)
            else:
                ws.cell(row=next_row, column=6).number_format = DATE_FORMAT
            next_row += 1
            rows_added += 1
        seen.add(batch.order_number)
        orders_added += 1

    if rows_added == 0:
        # Nothing new — leave the master completely untouched (no backup, no save).
        return None, 0, 0

    # Save a pristine copy of the master BEFORE overwriting it, as a restore point.
    backups_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = backups_dir / f"{master_xlsx.stem}_backup_{stamp}{master_xlsx.suffix}"
    shutil.copy2(master_xlsx, backup_path)

    wb.save(master_xlsx)
    return backup_path, rows_added, orders_added
